import csv
from schedule import every, repeat, run_pending
from time import sleep
from bs4 import BeautifulSoup
import requests
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy import (
    create_engine,
    MetaData,
    Table,
    Column,
    Integer,
    String,
    Float,
)

engine = create_engine("sqlite:///logs.db", echo=True)
meta = MetaData()

logs = Table(
    "logs",
    meta,
    Column("id", Integer, primary_key=True),
    Column("marka", String(80)),
    Column("branch", String(80)),
    Column("brand", String(80)),
    Column("date", String(80)),
    Column("hour", String(80)),
    Column("status", String(80)),
    Column("current_rating", Float),
)
meta.create_all(engine)


def save_to_db(marka, branch, brand, date, hour, status, current_rating):

    sql = logs.insert().values(
        marka=marka,
        branch=branch,
        brand=brand,
        date=date,
        hour=hour,
        status=status,
        current_rating=current_rating,
    )
    engine.execute(sql)


def save_to_db_log(marka, branch, brand, date, hour, log):

    sql = logs.insert().values(
        marka=marka, branch=branch, brand=brand, date=date, hour=hour, status=log, current_rating=log
    )


def save_to_excel(
    ws, wb, myFileName, marka, branch, brand, date, hour, status, current_rating
):
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=marka)
    ws.cell(row=1, column=2, value=branch)
    ws.cell(row=1, column=3, value=brand)
    ws.cell(row=1, column=4, value=date)
    ws.cell(row=1, column=5, value=hour)
    ws.cell(row=1, column=6, value=status)
    ws.cell(row=1, column=7, value=current_rating)

    # ws.append([branch, brand, date, hour, status, current_rating])

    wb.save(filename=myFileName)

    wb.close()


def save_to_excel_log(ws, wb, myFileName, marka, branch, brand, date, hour, log):


    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=marka)
    ws.cell(row=1, column=2, value=branch)
    ws.cell(row=1, column=3, value=brand)
    ws.cell(row=1, column=4, value=date)
    ws.cell(row=1, column=5, value=hour)
    ws.cell(row=1, column=6, value=log)
    ws.cell(row=1, column=7, value=log)

    # ws.append([branch, brand, date, hour, log, log])

    wb.save(filename=myFileName)

    wb.close()

def save_to_csv(branch,brand,date,hour,status,current_rating):
    None
def save_to_csv_log():
    None
def getir_to_excel_first():

    myFileName = r"./logs.xlsx"

    wb = load_workbook(filename=myFileName)

    # TODO: gonna create the page for every restaurant and make the changes into those restaurants
    ws = wb["Sheet1"]

    # TODO: Be sure that all the restaurant are open! And get the links. Then create a if statement to make OPEN/CLOSE diff. Belove statement is not quite right, gotta change it with the real time case.
    url_dict = {
        "Rakip Kadıköy Hyar Salad Bar": "https://getir.com/yemek/restoran/hyar-salad-bar-acibadem-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Green Salads": "https://getir.com/yemek/restoran/green-salads-fatih-cad-kadikoy-istanbul/",
        "Rakip Ataşehir By Yemekçi": "https://getir.com/yemek/restoran/by-yemekci-fetih-mah-atasehir-istanbul/",
        "Rakip Ümraniye Memo's Salad": "https://getir.com/yemek/restoran/memo-s-salad-tatlisu-mah-umraniye-istanbul/",
        "Rakip Ataşehir Abbas Gurme Salata": "https://getir.com/yemek/restoran/abbas-gurme-salata-yenisehir-mah-atasehir-istanbul/",
        "Rakip Ataşehir Brasserie Polonez Prime": "https://getir.com/yemek/restoran/abbas-gurme-salata-yenisehir-mah-atasehir-istanbul/",
        "Rakip Beşiktaş Lumos Healthy Kitchen Coffee": "https://getir.com/yemek/restoran/lumos-healty-kitchen-coffee-konaklar-mah-besiktas-istanbul/",
        "Rakip Sarıyer The Wrap": "https://getir.com/yemek/restoran/the-wrap-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Beşiktaş Green Salads": "https://getir.com/yemek/restoran/green-salads-kultur-mah-besiktas-istanbul/",
        "Rakip Sarıyer Baby Green Salad": "https://getir.com/yemek/restoran/baby-green-salad-bowl-wrap-sariyer-istanbul/",
        "Rakip Sarıyer The Wrap": "https://getir.com/yemek/restoran/the-wrap-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Sarıyer Zizou Express": "https://getir.com/yemek/restoran/zizou-express-maslak-mah-sariyer-istanbul/",
        "Rakip Ataşehir Salad Town": "https://getir.com/yemek/restoran/salad-town-barbaros-mah-atasehir-istanbul/",
        "Rakip Kağıthane Diyet 34": "https://getir.com/yemek/restoran/diyet34-kagithane-istanbul/",
        "Rakip Beşiktaş Hebun Çorba Evi": "https://getir.com/yemek/restoran/hebun-corba-evi-huzur-mah-besiktas-istanbul/",
        "Rakip Sariyer Suup": "https://getir.com/yemek/restoran/suup-corba-meze-resitpasa-mah-sariyer-istanbul/",
        "Rakip Şişli Salad to Queen": "https://getir.com/yemek/restoran/salad-to-queen-fulya-mah-sisli-istanbul/",
        "Rakip Şişli Salad Station": "https://getir.com/yemek/restoran/saladstation-esentepe-mah-sisli-istanbul/",
        "Rakip Kadıköy Alo Mantı": "https://getir.com/yemek/restoran/alo-manti-acibadem-mah-kadikoy-istanbul/",
        "Rakip Üsküdar Alesta Mantı": "https://getir.com/yemek/restoran/alesta-manti-ev-yemekleri-valide-i-atik-mah-uskudar-istanbul/",
        "Rakip Ataşehir Ferzande Mantı Evi": "https://getir.com/yemek/restoran/ferzande-manti-evi-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Mantı Dünyası": "https://getir.com/yemek/restoran/manti-dunyasi-atasehir-istanbul/",
        "Rakip Ataşehir Green Manti Evi": "https://getir.com/yemek/restoran/green-manti-evi-icerenkoy-mah-atasehir-istanbul/",
        "Rakip Beşiktaş Ulubey Manti": "https://getir.com/yemek/restoran/ulubey-manti-etiler-mah-besiktas-istanbul/",
        "Rakip Beşiktaş Casita Mantı": "https://getir.com/yemek/restoran/casita-manti-etiler-mah-besiktas-istanbul/",
        "Rakip Sariyer Manti Yedi": "https://getir.com/yemek/restoran/manti-yedi-sariyer-istanbul/",
        "Rakip Sariyer Askana Manti": "https://getir.com/yemek/restoran/askana-manti-maslak-mah-sariyer-istanbul/",
        "Rakip Sariyer Asude Manti": "https://getir.com/yemek/restoran/asude-manti-ayazaga-mah-sariyer-istanbul/",
        "Rakip Üsküdar Masal Ev Yemekleri": "https://getir.com/yemek/restoran/masal-ev-yemekleri-cafe-selimiye-mah-uskudar-istanbul/",
        "Rakip Üsküdar Portakal Çiçeği Mutfağı": "https://getir.com/yemek/restoran/portakal-cicegi-mutfagi-acibadem-mah-uskudar-istanbul/",
        "Rakip Üsküdar Divan Delivery": "https://getir.com/yemek/restoran/divan-delivery-altunizade-mah-uskudar-istanbul-2/",
        "Rakip Ataşehir Divan Delivery": "https://getir.com/yemek/restoran/divan-delivery-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Çorba İstanbul": "https://getir.com/yemek/restoran/corba-istanbul-et-mangal-atasehir-istanbul/",
        "Rakip Beşiktaş Hanım Eli": "https://getir.com/yemek/restoran/hanim-eli-dikilitas-mah-besiktas-istanbul/",
        "Rakip Beşiktaş Kozlet": "https://getir.com/yemek/restoran/kozlet-besiktas-istanbul/",
        "Rakip Sarıyer Soft Mutfak": "https://getir.com/yemek/restoran/soft-mutfak-istinye-mah-sariyer-istanbul/",
        "Rakip Kadıköy Veganarsist": "https://getir.com/yemek/restoran/veganarsist-osmanaga-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Limonita Vegan Mutfak": "https://getir.com/yemek/restoran/limonita-vegan-mutfak-caferaga-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Vatka Coffee Vegan Goods": "https://getir.com/yemek/restoran/vatka-coffee-vegan-goods-caferaga-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Zebze": "https://getir.com/yemek/restoran/zebze-yenisehir-mah-atasehir-istanbul/",
        "Rakip Ataşehir Let's Salad": "https://getir.com/yemek/restoran/let-s-salad-barbaros-mah-atasehir-istanbul/",
        "Rakip Ataşehir Salad Town": "https://getir.com/yemek/restoran/salad-town-barbaros-mah-atasehir-istanbul/",
        "Rakip Beşiktaş Vegan Street Food": "https://getir.com/yemek/restoran/vegan-street-food-bebek-mah-besiktas-istanbul/",
        "Rakip Beşiktaş Mr Bean Vegan": "https://getir.com/yemek/restoran/mr-bean-vegan-nisbetiye-mah-besiktas-istanbul/",
        "Rakip Şişli Govinda İstanbul": "https://getir.com/yemek/restoran/govinda-istanbul-mecidiyekoy-mah-sisli-istanbul/",
        "Rakip Sarıyer Salad Town": "https://getir.com/yemek/restoran/salad-town-maslak-mah-sariyer-istanbul/",
        "Rakip Sarıyer Let's Salad": "https://getir.com/yemek/restoran/let-s-salad-maslak-mah-sariyer-istanbul/",
        "Rakip Ataşehir Pizza Lazza": "https://getir.com/yemek/restoran/pizza-lazza-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Dominos Pizza": "https://getir.com/yemek/restoran/domino-s-pizza-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Pizza Da Villa": "https://getir.com/yemek/restoran/pizza-da-villa-ataturk-mah-atasehir-istanbul/",
        "Rakip Ümraniye Chicken Burger Grill": "https://getir.com/yemek/restoran/chicken-burger-grill-mehmet-akif-mah-umraniye-istanbul/",
        "Rakip Kadıköy Albero": "https://getir.com/yemek/restoran/albero-kozyatagi-mah-kadikoy-istanbul/",
        "Rakip Ümraniye MdMad Buger": "https://getir.com/yemek/restoran/mdmad-burger-serifali-mah-umraniye-istanbul/",
        "Rakip Beşiktaş Popeyes": "https://getir.com/yemek/restoran/popeyes-besiktas-etiler-mah-besiktas-istanbul/",
        "Rakip Sarıyer Cross Fingers": "https://getir.com/yemek/restoran/cross-fingers-etiler-mah-sariyer-istanbul/",
        "Rakip Sarıyer Rebel Street Food": "https://getir.com/yemek/restoran/rebel-street-food-resitpasa-mah-sariyer-istanbul/",
        "Rakip Sarıyer McDonalds": "https://getir.com/yemek/restoran/mcdonald-s-maslak-mah-sariyer-istanbul/",
        "Rakip Sarıyer KFC": "https://getir.com/yemek/restoran/kfc-maslak-mah-sariyer-istanbul/",
        "Rakip Sarıyer Carls JR": "https://getir.com/yemek/restoran/carl-s-jr-ayazaga-mah-sariyer-istanbul/",
        "Rakip Kadıköy Rolla": "https://getir.com/yemek/restoran/rolla-kadikoy-istanbul/",
        "Rakip Kadıköy 700 Gram": "https://getir.com/yemek/restoran/700-gram-caferaga-mah-kadikoy-istanbul/",
        "Rakip Beşiktaş Overdose": "https://getir.com/yemek/restoran/overdose-istanbul-besiktas-istanbul/",
        "Rakip Kağıthane Diet Palace": "https://getir.com/yemek/restoran/diet-palace-emniyetevleri-mah-kagithane-istanbul/",
        "Rakip Sarıyer Formist Healthy Foods": "https://getir.com/yemek/restoran/formist-healthy-foods-maslak-mah-sariyer-istanbul/",
        "Rakip Sarıyer Fit Chef": "https://getir.com/yemek/restoran/fit-chef-pinar-mah-sariyer-istanbul/",
        "Rakip Sarıyer Fit ve Hafif": "https://getir.com/yemek/restoran/fit-ve-hafif-maslak-mah-sariyer-istanbul/",
        "Rakip Kadıköy Meze Roll": "https://getir.com/yemek/restoran/meze-roll-acibadem-mah-kadikoy-istanbul/",
        "Rakip Kadıköy İkinci Meze Zeytinyağlı": "https://getir.com/yemek/restoran/ikinci-meze-zeytinyagli-caferaga-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Acıbadem Meze Evi": "https://getir.com/yemek/restoran/acibadem-meze-evi-acibadem-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Serbele Meze": "https://getir.com/yemek/restoran/serbele-meze-sahrayicedid-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Mezeci Melahat": "https://getir.com/yemek/restoran/mezeci-melahat-barbaros-mah-atasehir-istanbul/",
        "Rakip Ataşehir Meze Center": "https://getir.com/yemek/restoran/meze-center-ornek-mah-atasehir-istanbul/",
        "Rakip Sariyer Mezepoly": "https://getir.com/yemek/restoran/mezepoly-fatih-sultan-mehmet-mah-sariyer-istanbul-2/",
        "Rakip Sariyer Yan Mutfak": "https://getir.com/yemek/restoran/yan-mutfak-sariyer-istanbul/",
        "Rakip Ataşehir KFC": "https://getir.com/yemek/restoran/kfc-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Popeyes": "https://getir.com/yemek/restoran/popeyes-atasehir-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir McDonalds": "https://getir.com/yemek/restoran/mcdonald-s-atasehir-merkez-atasehir-istanbul/",
        "Rakip Beşiktaş Popeyes": "https://getir.com/yemek/restoran/popeyes-besiktas-etiler-mah-besiktas-istanbul/",
        "Rakip Sariyer Rebel Street Food": "https://getir.com/yemek/restoran/rebel-street-food-resitpasa-mah-sariyer-istanbul/",
        "Rakip Sariyer McDonalds": "https://getir.com/yemek/restoran/mcdonald-s-maslak-mah-sariyer-istanbul/",
        "Rakip Sariyer Carls Jr": "https://getir.com/yemek/restoran/carl-s-jr-ayazaga-mah-sariyer-istanbul/",
        "Rakip Ataşehir Bowl More": "https://getir.com/yemek/restoran/bowl-more-barbaros-mah-atasehir-istanbul/",
        "Rakip Ataşehir Cookshop": "https://getir.com/yemek/restoran/cookshop-ataturk-mah-atasehir-istanbul/",
        "Rakip Beşiktaş Overdose ": "https://getir.com/yemek/restoran/overdose-istanbul-besiktas-istanbul/",
        "Rakip Beşiktaş Bowl Department": "https://getir.com/yemek/restoran/bowl-department-akat-mah-besiktas-istanbul/",
        "Rakip Sariyer Gina Bowl": "https://getir.com/yemek/restoran/gina-bowl-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Sariyer Monk": "https://getir.com/yemek/restoran/monk-huzur-mah-sariyer-istanbul/",
        "Rakip Sariyer Baby Green Salad Bowl Wrap": "https://getir.com/yemek/restoran/baby-green-salad-bowl-wrap-sariyer-istanbul/",
        "Rakip Sariyer Bowl More": "https://getir.com/yemek/restoran/bowl-more-maslak-mah-sariyer-istanbul/",
        "Rakip Ataşehir Tikka Tavuk": "https://getir.com/yemek/restoran/tikka-tavuk-kucukbakkalkoy-mah-atasehir-istanbul/",
        "Rakip Ataşehir Maydonoz Döner": "https://getir.com/yemek/restoran/maydonoz-doner-kucukbakkalkoy-mah-atasehir-istanbul/",
        "Rakip Ataşehir Hot Döner": "https://getir.com/yemek/restoran/hot-doner-kayisdagi-mah-atasehir-istanbul/",
        "Rakip Sariyer Dönerci Cihan Usta": "https://getir.com/yemek/restoran/donerci-cihan-usta-fsm-mah-sariyer-istanbul/",
        "Rakip Sariyer Hot Döner": "https://getir.com/yemek/restoran/hot-doner-rumelihisari-mah-sariyer-istanbul/",
        "Rakip Kağıthane Ali Baba Döner": "https://getir.com/yemek/restoran/ali-baba-doner-celiktepe-mah-kagithane-istanbul/",
        "Rakip Sariyer İkram Döner": "https://getir.com/yemek/restoran/ikram-doner-ayazaga-mah-sariyer-istanbul/",
        "Rakip Sariyer Paşa Döner": "https://getir.com/yemek/restoran/pasa-doner-hisarustu-mah-sariyer-istanbul/",
        "Rakip Üsküdar 6 Üstü Krık Köfte": "https://getir.com/yemek/restoran/6-ustu-kirk-kofte-mimar-sinan-mah-uskudar-istanbul/",
        "Rakip Ümraniye Şerifali Köftecisi": "https://getir.com/yemek/restoran/serifali-koftecisi-serifali-mah-umraniye-istanbul/",
        "Rakip Ataşehir Vefalı Köfteci": "https://getir.com/yemek/restoran/vefali-kofteci-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Abbas Köfteci": "https://getir.com/yemek/restoran/abbas-kofteci-ataturk-mah-atasehir-istanbul/",
        "Rakip Sariyer Mecburiyet Köftecisi": "https://getir.com/yemek/restoran/mecburiyet-koftecisi-maslak-mah-sariyer-istanbul/",
        "Rakip Sariyer Bi Ton Köfte": "https://getir.com/yemek/restoran/bi-ton-kofte-resitpasa-mah-sariyer-istanbul/",
        "Rakip Beşiktaş Günaydın Köfte Döner": "https://getir.com/yemek/restoran/gunaydin-kofte-doner-nisbetiye-mah-besiktas-istanbul/",
        "Rakip Şişli Köftehane": "https://getir.com/yemek/restoran/koftehane-sisli-istanbul/",
        "Rakip Sariyer Etkolik": "https://getir.com/yemek/restoran/etkolik-maslak-mah-sariyer-istanbul/",
        "Rakip Sariyer Bi Ton Köfte": "https://getir.com/yemek/restoran/bi-ton-kofte-resitpasa-mah-sariyer-istanbul/",
        "Rakip Kadıköy No 90 Vintage Coffee Burger": "https://getir.com/yemek/restoran/no-90-vintage-coffee-burger-kadikoy-istanbul/",
        "Rakip Kadıköy Falafella": "https://getir.com/yemek/restoran/falafella-caferaga-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Sugo Street Food": "https://getir.com/yemek/restoran/sugo-street-food-barbaros-mah-atasehir-istanbul/",
        "Rakip Ümraniye Burger End": "https://getir.com/yemek/restoran/burger-end-serifali-mah-umraniye-istanbul/",
        "Rakip Ümraniye Rivella Cafe": "https://getir.com/yemek/restoran/rivella-cafe-restaurant-serifali-mah-umraniye-istanbul/",
        "Rakip Sariyer Büfe Bu Gurme": "https://getir.com/yemek/restoran/bufe-bu-gurme-rumelihisari-mah-sariyer-istanbul/",
        "Rakip Beşiktaş Vegan Street Food": "https://getir.com/yemek/restoran/vegan-street-food-bebek-mah-besiktas-istanbul/",
        "Rakip Sariyer Leyna Falafel": "https://getir.com/yemek/restoran/leyna-falafel-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Sariyer Fanfan Cafe": "https://getir.com/yemek/restoran/fanfan-cafe-resitpasa-mah-sariyer-istanbul/",
        "Rakip Sariyer Meat Burger Gurme Mutfak": "https://getir.com/yemek/restoran/meat-burger-gurme-mutfak-maslak-mah-sariyer-istanbul/",
        "Rakip Ataşehir Sushi G": "https://getir.com/yemek/restoran/sushi-g-barbaros-mah-atasehir-istanbul/",
        "Rakip Ataşehir Chinese Sushi Express": "https://getir.com/yemek/restoran/chinese-sushi-express-barbaros-mah-atasehir-istanbul/",
        "Rakip Ataşehir Sushi Manga": "https://getir.com/yemek/restoran/sushi-manga-barbaros-mah-atasehir-istanbul/",
        "Rakip Kadıköy 30 06 Springfield": "https://getir.com/yemek/restoran/30-06-springfield-bostanci-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Frango Döner Sauce": "https://getir.com/yemek/restoran/frango-doner-sauce-feneryolu-mah-kadikoy-istanbul/",
        "Rakip Kadıköy KFC": "https://getir.com/yemek/restoran/kfc-acibadem-mah-kadikoy-istanbul/",
        "Rakip Ataşehir KFC": "https://getir.com/yemek/restoran/kfc-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Popeyes": "https://getir.com/yemek/restoran/popeyes-atasehir-ataturk-mah-atasehir-istanbul/",
        "Rakip Kadıköy Albero": "https://getir.com/yemek/restoran/albero-kozyatagi-mah-kadikoy-istanbul/",
        "Rakip Beşiktaş Popeyes": "https://getir.com/yemek/restoran/popeyes-besiktas-etiler-mah-besiktas-istanbul/",
        "Rakip Sariyer El Torito Mexican": "https://getir.com/yemek/restoran/el-torito-mexican-etiler-mah-sariyer-istanbul/",
        "Rakip Sariyer Wrapetito": "https://getir.com/yemek/restoran/wrapetito-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Sariyer The Wrap": "https://getir.com/yemek/restoran/the-wrap-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Üsküdar Burgerillas": "https://getir.com/yemek/restoran/burgerillas-acibadem-mah-uskudar-istanbul/",
        "Rakip Kadıköy Brand Burger": "https://getir.com/yemek/restoran/brand-burger-acibadem-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Burgerist": "https://getir.com/yemek/restoran/burgerist-acibadem-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Burgerillas": "https://getir.com/yemek/restoran/burgerillas-ataturk-mah-atasehir-istanbul/",
        "Rakip Ümraniye Memo's Burger Hot Dog": "https://getir.com/yemek/restoran/memo-s-burger-hot-dog-tatlisu-mah-umraniye-istanbul/",
        "Rakip Sariyer Vitto Mama": "https://getir.com/yemek/restoran/vitto-mama-barbaros-mah-atasehir-istanbul/",
        "Rakip Sariyer Nera Burger": "https://getir.com/yemek/restoran/nera-burger-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Beşiktaş Nişantaşı Burger Chicken": "https://getir.com/yemek/restoran/nisantasi-burger-chicken-akat-mah-besiktas-istanbul/",
        "Rakip Beşiktaş Burgerillas": "https://getir.com/yemek/restoran/burgerillas-nisbetiye-mah-besiktas-istanbul-2/",
        "Rakip Kağıthane Burgerzoom": "https://getir.com/yemek/restoran/burgerzoom-merkez-mah-kagithane-istanbul/",
        "Rakip Kağıthane Kiraz Burger": "https://getir.com/yemek/restoran/kiraz-burger-sultan-selim-mah-kagithane-istanbul/",
        "Rakip Sariyer Boots Spurs Burger": "https://getir.com/yemek/restoran/boots-spurs-burger-ayazaga-mah-sariyer-istanbul/",
        "Rakip Ataşehir Duppo": "https://getir.com/yemek/restoran/duppo-atasehir-istanbul/",
        "Rakip Ataşehir Sugar Cheese": "https://getir.com/yemek/restoran/sugar-cheese-ataturk-mah-atasehir-istanbul/",
        "Rakip Beşiktaş My Sufle Dondurma": "https://getir.com/yemek/restoran/my-sufle-dondurma-etiler-mah-besiktas-istanbul/",
        "Rakip Beşiktaş Süflor Dondurma": "https://getir.com/yemek/restoran/suflor-dondurma-akat-mah-besiktas-istanbul/",
        "Rakip Kağıthane Hero's Tatlı": "https://getir.com/yemek/restoran/hero-s-tatli-hamidiye-mah-kagithane-istanbul-2/",
        "Rakip Kadıköy No 90 Coffee Burger": "https://getir.com/yemek/restoran/no-90-vintage-coffee-burger-kadikoy-istanbul/",
        "Rakip Ataşehir Chick Billy": "https://getir.com/yemek/restoran/chick-billy-barbaros-mah-atasehir-istanbul/",
        "Rakip Ümraniye Rivella Cafe": "https://getir.com/yemek/restoran/rivella-cafe-restaurant-serifali-mah-umraniye-istanbul/",
        "Rakip Ataşehir Sugo Street Food": "https://getir.com/yemek/restoran/sugo-street-food-barbaros-mah-atasehir-istanbul/",
        "Rakip Kadıköy Albero": "https://getir.com/yemek/restoran/albero-kozyatagi-mah-kadikoy-istanbul/",
        "Rakip Sariyer Chop-T": "https://getir.com/yemek/restoran/chop-t-maslak-mah-sariyer-istanbul/",
        "Rakip Beşiktaş Vegan Street Food": "https://getir.com/yemek/restoran/vegan-street-food-bebek-mah-besiktas-istanbul/",
        "Rakip Sariyer FanFan Cafe": "https://getir.com/yemek/restoran/fanfan-cafe-resitpasa-mah-sariyer-istanbul/",
        "Rakip Sariyer HopDaddy Burger": "https://getir.com/yemek/restoran/hopdaddy-burger-sariyer-istanbul/",
        "Rakip Kağıthane Burgerzoom": "https://getir.com/yemek/restoran/burgerzoom-merkez-mah-kagithane-istanbul/",
        "Rakip Ataşehir Guru": "https://getir.com/yemek/restoran/guru-barbaros-mah-atasehir-istanbul/",
        "Rakip Ataşehir Sugar Cheese": "https://getir.com/yemek/restoran/sugar-cheese-ataturk-mah-atasehir-istanbul/",
        "Rakip Sarıyer Espumoso Kahve Tatlı": "https://getir.com/yemek/restoran/espumoso-kahve-tatli-rumelihisari-mah-sariyer-istanbul/",
        "Rakip Beşiktaş Kahve Dünyası": "https://getir.com/yemek/restoran/kahve-dunyasi-levent-carsi-mah-besiktas-istanbul/",
        "Rakip Sariyer Azelia Patisserie": "https://getir.com/yemek/restoran/azelia-patisserie-tarabya-mah-sariyer-istanbul-2/",
        "Rakip Üsküdar Big Chefs": "https://getir.com/yemek/restoran/big-chefs-altunizade-mah-uskudar-istanbul/",
        "Rakip Kadıköy Bams Cafe": "https://getir.com/yemek/restoran/bam-s-cafe-restoran-acibadem-mah-kadikoy-istanbul/",
        "Rakip Üsküdar Sera House Cafe": "https://getir.com/yemek/restoran/sera-house-cafe-uskudar-istanbul/",
        "Rakip Ataşehir Ateş Kitchen Coffee": "https://getir.com/yemek/restoran/ates-kitchen-coffee-kayisdagi-mah-atasehir-istanbul/",
        "Rakip Sariyer Tatties Etiler": "https://getir.com/yemek/restoran/tatties-etiler-rumelihisari-mah-sariyer-istanbul/",
        "Rakip Sariyer The Concrete House Cafe": "https://getir.com/yemek/restoran/the-concrete-house-cafe-etiler-mah-sariyer-istanbul/",
        "Rakip Kağıthane Quqi Cafe Patisserie": "https://getir.com/yemek/restoran/quqi-cafe-patisserie-sultan-selim-mah-kagithane-istanbul/",
        "Rakip Sariyer Mikado Cafe": "https://getir.com/yemek/restoran/mikado-cafe-restaurant-ayazaga-mah-sariyer-istanbul/",
        "Rakip Sariyer BKM Mutfak Cafe": "https://getir.com/yemek/restoran/bkm-mutfak-cafe-restaurant-sariyer-istanbul/",
        "Rakip Kadıköy Dürümle": "https://getir.com/yemek/restoran/durumle-acibadem-mah-nautilus-avm-kadikoy-istanbul/",
        "Rakip Kadıköy Emek Dürüm": "https://getir.com/yemek/restoran/emek-durum-kadikoy-istanbul/",
        "Rakip Ataşehir Dragon Selective Tastes": "https://getir.com/yemek/restoran/dragon-selective-tastes-ataturk-mah-atasehir-istanbul/",
        "Rakip Sariyer El Torito Mexican": "https://getir.com/yemek/restoran/el-torito-mexican-etiler-mah-sariyer-istanbul/",
        "Rakip Sariyer Wrapetito": "https://getir.com/yemek/restoran/wrapetito-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Rakip Sariyer Don Carlos Taco": "https://getir.com/yemek/restoran/don-carlos-taco-rumelihisari-mah-sariyer-istanbul/",
        "Rakip Sariyer Pack Ala Tacos Burritos Burgers": "https://getir.com/yemek/restoran/pack-ala-tacos-burritos-burgers-maslak-mah-sariyer-istanbul/",
        "Rakip Sariyer The Green Box": "https://getir.com/yemek/restoran/the-green-box-sariyer-istanbul/",
        "Rakip Sariyer Chop T": "https://getir.com/yemek/restoran/chop-t-maslak-mah-sariyer-istanbul/",
        "Rakip Kadıköy Archibalds Salad House": "https://getir.com/yemek/restoran/archibalds-salad-house-zuhtupasa-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Office Gastro Salad": "https://getir.com/yemek/restoran/office-gastro-salad-barbaros-mah-atasehir-istanbul/",
        "Rakip Beşiktaş Green Queen Salads": "https://getir.com/yemek/restoran/green-queen-salads-levent-mah-besiktas-istanbul/",
        "Rakip Beşiktaş Green Salads": "https://getir.com/yemek/restoran/green-salads-kultur-mah-besiktas-istanbul/",
        "Rakip Kadıköy Tavuk Dünyası": "https://getir.com/yemek/restoran/tavuk-dunyasi-osmanaga-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Green Salads": "https://getir.com/yemek/restoran/green-salads-fatih-cad-kadikoy-istanbul/",
        "Rakip Ataşehir Green Salads": "https://getir.com/yemek/restoran/green-salads-ataturk-mah-atasehir-istanbul/",
        "Rakip Ümraniye Simit Tadında": "https://getir.com/yemek/restoran/simit-tadinda-umraniye-istanbul/",
        "Rakip Ümraniye Park Piramit": "https://getir.com/yemek/restoran/park-piramit-camlik-mah-umraniye-istanbul/",
        "Rakip Beşiktaş Green Salads": "https://getir.com/yemek/restoran/green-salads-kultur-mah-besiktas-istanbul/",
        "Rakip Sarıyer Food Garden": "https://getir.com/yemek/restoran/food-garden-pinar-mah-sariyer-istanbul/",
        "Rakip Sarıyer The Concrete House Cafe": "https://getir.com/yemek/restoran/the-concrete-house-cafe-etiler-mah-sariyer-istanbul/",
        "Rakip Kadıköy Murat Muhallebi": "https://getir.com/yemek/restoran/murat-muhallebi-rasimpasa-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Moda": "https://getir.com/yemek/restoran/mado-rasimpasa-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Zeynel Muhallebicisi": "https://getir.com/yemek/restoran/zeynel-muhallebicisi-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Kadıköy Saray Muhallebicisi": "https://getir.com/yemek/restoran/kadikoy-saray-muhallebicisi-ataturk-mah-atasehir-istanbul/",
        "Rakip Kadıköy Murat Muhallebi": "https://getir.com/yemek/restoran/murat-muhallebi-sahrayicedit-mah-kadikoy-istanbul/",
        "Rakip Sarıyer Kaçkar Muhallebicisi": "https://getir.com/yemek/restoran/kackar-muhallebicisi-sariyer-istanbul/",
        "Rakip Beşiktaş Bolulu Hasan Usta": "https://getir.com/yemek/restoran/bolulu-hasan-usta-nisbetiye-mah-besiktas-istanbul/",
        "Rakip Sariyer Zeynel Muhallebicisi": "https://getir.com/yemek/restoran/zeynel-muhallebicisi-yenikoy-mah-sariyer-istanbul/",
        "Rakip Şişli Bal Badem Pastanesi": "https://getir.com/yemek/restoran/balbadem-pastanesi-sisli-istanbul/",
        "Rakip Kadıköy Mis Tantuni": "https://getir.com/yemek/restoran/mis-tantuni-caferaga-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Mr Tantuni Sahrayıcedit": "https://getir.com/yemek/restoran/mr-tantuni-sahrayicedit-mah-kadikoy-istanbul/",
        "Rakip Kadiköy Mert 33 Tantuni": "https://getir.com/yemek/restoran/mert-33-tantuni-kadikoy-istanbul/",
        "Rakip Kadıköy Fırat Dürüm": "https://getir.com/yemek/restoran/firat-durum-merdivenkoy-mah-kadikoy-istanbul/",
        "Rakip Üsküdar Sıfır Bir Adanalı Dürümcü": "https://getir.com/yemek/restoran/sifirbir-adanali-durumcu-icadiye-mah-uskudar-istanbul/",
        "Rakip Ataşehir Makarnam Fratelli": "https://getir.com/yemek/restoran/makarnam-fratelli-ataturk-mah-atasehir-istanbul/",
        "Rakip Ataşehir Mektep Dürüm": "https://getir.com/yemek/restoran/mektep-durum-kayisdagi-mah-atasehir-istanbul/",
        "Rakip Kağıthane Dürümcü Onur Usta'nin Yeri": "https://getir.com/yemek/restoran/durumcu-onur-usta-nin-yeri-ortabayir-mah-kagithane-istanbul/",
        "Rakip Sariyer Dürümcü Nevzat Usta": "https://getir.com/yemek/restoran/durumcu-nevzat-usta-ayazaga-mah-sariyer-istanbul/",
        "Rakip Sariyer Bizim Dürümcü": "https://getir.com/yemek/restoran/bizim-durumcu-cumhuriyet-mah-sariyer-istanbul/",
        "Rakip Kadıköy Paul s Lasagna": "https://getir.com/yemek/restoran/paul-s-lasagna-rasimpasa-mah-kadikoy-istanbul/",
        "Rakip Üsküdar Gurme Mutfak": "https://getir.com/yemek/restoran/gurme-mutfak-altunizade-mah-uskudar-istanbul/",
        "Rakip Kadıköy Kaen Sushi": "https://getir.com/yemek/restoran/kaen-sushi-kosuyolu-mah-kadikoy-istanbul/",
        "Rakip Üsküdar Sushico": "https://getir.com/yemek/restoran/sushico-acibadem-mah-uskudar-istanbul/",
        "Rakip Üsküdar Chinese Sushi Express": "https://getir.com/yemek/restoran/chinese-sushi-express-altunizade-mah-uskudar-istanbul/",
        "Rakip Kadıköy Akveren Makarna": "https://getir.com/yemek/restoran/akveren-makarna-kadikoy-istanbul/",
        "Rakip Kadıköy Gurme Mutfak": "https://getir.com/yemek/restoran/gurme-mutfak-acibadem-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Kaseden Makarna Salata": "https://getir.com/yemek/restoran/kaseden-makarna-salata-yenisahra-mah-atasehir-istanbul/",
        "Rakip Ataşehir Makarnam Fratelli": "https://getir.com/yemek/restoran/makarnam-fratelli-ataturk-mah-atasehir-istanbul/",
        "Rakip Sarıyer Cafe Boyacıköy": "https://getir.com/yemek/restoran/cafe-boyacikoy-sariyer-istanbul/",
        "Rakip Kağıthane Gurme Mutfak": "https://getir.com/yemek/restoran/gurme-mutfak-emniyet-evleri-mah-kagithane-istanbul/",
        "Rakip Sariyer Makarna Company": "https://getir.com/yemek/restoran/makarna-company-maslak-mah-sariyer-istanbul/",
        "Rakip Kağıthane Başka Makarna": "https://getir.com/yemek/restoran/baska-makarna-yesilce-mah-kagithane-istanbul/",
        "Rakip Ataşehir Baja Azteca": "https://getir.com/yemek/restoran/baja-azteca-barbaros-mah-atasehir-istanbul/",
        "Rakip Kadıköy Burrito Shop": "https://getir.com/yemek/restoran/burrito-shop-caferaga-mah-kadikoy-istanbul/",
        "Rakip Kadıköy Alfred Cafe Restaurant": "https://getir.com/yemek/restoran/alfred-cafe-restaurant-caferaga-mah-kadikoy-istanbul/",
        "Rakip Ataşehir Tacofit": "https://getir.com/yemek/restoran/tacofit-ataturk-mah-atasehir-istanbul/",
        "Avane Acıbadem Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Çosa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Doyuyo": "https://getir.com/yemek/restoran/doyuyo-sarayardi-cad-kadikoy-istanbul/",
        "Avane Acıbadem El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Etişler Köfte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Fadelini": "https://getir.com/yemek/restoran/fadelini-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-acibadem-mah-kadikoy-istanbul-2/",
        "Avane Acıbadem Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Kale Arkası Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Kengeres Çiğ Köfte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Madritas": "https://getir.com/yemek/restoran/madritas-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Nane Mantı": "https://getir.com/yemek/restoran/nane-manti-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Noody": "https://getir.com/yemek/restoran/noody-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Seez Beez": "https://getir.com/yemek/restoran/seez-beez-falafel-wraps-burgers-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-acibadem-mah-kadikoy-istanbul/",
        "Avane Acıbadem Tabur Köfte": "https://getir.com/yemek/restoran/tabur-kofte-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-acibadem-mah-kadikoy-istanbul",
        "Avane Acıbadem Veganista": "https://getir.com/yemek/restoran/veganista-acibadem-mah-kadikoy-istanbul",
        "Avane Ataşehir Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Çosa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Doyuyo": "https://getir.com/yemek/restoran/doyuyo-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Fadelini": "https://getir.com/yemek/restoran/fadelini-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Fun For Fit": " https://getir.com/yemek/restoran/fun-for-fit-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Kale Arkası Mutfak": " https://getir.com/yemek/restoran/kale-arkasi-mutfak-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Kengeres Çiğ Köfte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Madritas": "https://getir.com/yemek/restoran/madritas-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Mztps Meze": " https://getir.com/yemek/restoran/mztps-meze-tapas-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Nane Mantı": " https://getir.com/yemek/restoran/nane-manti-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Noody": "https://getir.com/yemek/restoran/noody-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Seez Beez": "https://getir.com/yemek/restoran/seez-beez-falafel-wraps-burgers-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir Sushi Master": "https://getir.com/yemek/restoran/sushi-master-ataturk-mah-atasehir-istanbul/",
        "Avane Ataşehir The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-ataturk-mah-atasehir-istanbul",
        "Avane Ataşehir Veganista": "https://getir.com/yemek/restoran/veganista-ataturk-mah-atasehir-istanbul",
        "Avane Kozyatağı Ali Veli Gurme Pide": "https://getir.com/yemek/restoran/ali-veli-gurme-pide-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Çosa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Doyuyo": "https://getir.com/yemek/restoran/doyuyo-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-kozyatagi-mah-kadikoy-istanbul//",
        "Avane Kozyatağı Etişler Köfte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Fadelini": "https://getir.com/yemek/restoran/fadelini-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Kale Arkası Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Kengeres Çiğ Köfte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Madritas": "https://getir.com/yemek/restoran/madritas-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Nane Mantı": "https://getir.com/yemek/restoran/nane-manti-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Noody": "https://getir.com/yemek/restoran/noody-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Sushi Master": "https://getir.com/yemek/restoran/sushi-master-kozyatagi-mah-kadikoy-istanbul/",
        "Avane Kozyatağı Tabur Köfte": "https://getir.com/yemek/restoran/tabur-kofte-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-kozyatagi-mah-kadikoy-istanbul",
        "Avane Kozyatağı Veganista": "https://getir.com/yemek/restoran/veganista-kozyatagi-mah-kadikoy-istanbul",
        "Avane FSM Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM Çosa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Etişler Köfte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Fadelini": "https://getir.com/yemek/restoran/fadelini-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM Kale Arkası Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM Kengeres Çiğ Köfte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM Madritas": "https://getir.com/yemek/restoran/madritas-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-kozyatagi-mah-kadikoy-istanbul",
        "Avane FSM Nane Mantı": "https://getir.com/yemek/restoran/nane-manti-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Noody": "https://getir.com/yemek/restoran/noody-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Sushi Master": "https://getir.com/yemek/restoran/sushi-master-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM Tabur Köfte": "https://getir.com/yemek/restoran/tabur-kofte-fatih-sultan-mehmet-mah-sariyer-istanbul/",
        "Avane FSM The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane FSM Veganista": "https://getir.com/yemek/restoran/veganista-fatih-sultan-mehmet-mah-sariyer-istanbul",
        "Avane Maslak Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-maslak-mah-sariyer-istanbul",
        "Avane Maslak Çosa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Detroit Bad Boys Pizza": "https://getir.com/yemek/restoran/detroit-bad-boys-pizza-maslak-mah-sariyer-istanbul",
        "Avane Maslak Dlycious Dyssert": "https://getir.com/yemek/restoran/dydy-dylicious-dyssert-maslak-mah-sariyer-istanbul",
        "Avane Maslak El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Etişler Köfte": "https://getir.com/yemek/restoran/et-isleri-kofte-burger-durum-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Fadelini": "https://getir.com/yemek/restoran/fadelini-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-maslak-mah-sariyer-istanbul",
        "Avane Maslak G&G Burger": "https://getir.com/yemek/restoran/g-g-burger-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-maslak-mah-sariyer-istanbul",
        "Avane Maslak Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-maslak-mah-sariyer-istanbul",
        "Avane Maslak Kale Arkası Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-maslak-mah-sariyer-istanbul",
        "Avane Maslak Kengeres Çiğ Köfte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-maslak-mah-sariyer-istanbul",
        "Avane Maslak Madritas": "https://getir.com/yemek/restoran/madritas-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Mztps Meze": "https://getir.com/yemek/restoran/mztps-meze-tapas-maslak-mah-sariyer-istanbul",
        "Avane Maslak Nane Mantı": "https://getir.com/yemek/restoran/nane-manti-evi-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Noody": "https://getir.com/yemek/restoran/noody-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Red Haag": "https://getir.com/yemek/restoran/red-haag-cafe-brasserie-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Rylee's Ranch Salad": "https://getir.com/yemek/restoran/rylee-s-ranch-salad-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Sushi Master": "https://getir.com/yemek/restoran/sushi-master-maslak-mah-sariyer-istanbul/",
        "Avane Maslak Tabur Köfte": "https://getir.com/yemek/restoran/tabur-kofte-maslak-mah-sariyer-istanbul/",
        "Avane Maslak The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-maslak-mah-sariyer-istanbul",
        "Avane Maslak Veganista": "https://getir.com/yemek/restoran/veganista-maslak-mah-sariyer-istanbul",
        "Avane İzmir Alle Bowls": "https://getir.com/yemek/restoran/alle-bowls-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir Arianas Cheesecake": "https://getir.com/yemek/restoran/ariana-s-cheesecake-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir Big Bold Quick": "https://getir.com/yemek/restoran/bbq-big-bold-quick-atasehir-mah-cigli-izmir/",
        "Avane İzmir Caesar Salad By": "https://getir.com/yemek/restoran/caesar-salad-by-chef-amadeo-cigli-atasehir-mah-cigli-izmir",
        "Avane İzmir Çosa": "https://getir.com/yemek/restoran/cosa-bi-corba-bi-salata-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir El Pollo Lasso": "https://getir.com/yemek/restoran/el-pollo-lasso-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir Etişler Köfte": "https://getir.com/yemek/restoran/etisleri-kofte-burger-durum-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir Fadelini": "https://getir.com/yemek/restoran/fadelini-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir Fun For Fit": "https://getir.com/yemek/restoran/fun-for-fit-maslak-mah-sariyer-istanbul",
        "Avane İzmir Gurra Tavuk": "https://getir.com/yemek/restoran/gurra-tavuk-doner-cigli-atasehir-mah-cigli-izmir",
        "Avane İzmir Jay Jay Fries": "https://getir.com/yemek/restoran/jay-jay-fries-maslak-mah-sariyer-istanbul",
        "Avane İzmir Kale Arkası Mutfak": "https://getir.com/yemek/restoran/kale-arkasi-mutfak-cigli-atasehir-mah-cigli-izmir",
        "Avane İzmir Kengeres Çiğ Köfte": "https://getir.com/yemek/restoran/kengeres-gurme-cig-kofte-cigli-atasehir-mah-cigli-izmir",
        "Avane İzmir Madritas": "https://getir.com/yemek/restoran/madritas-cigli-atasl-burgers-wraps-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir Senor Torreon": "https://getir.com/yemek/restoran/senor-torreon-red-hot-spicy-food-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir Tabur Köfte": "https://getir.com/yemek/restoran/tabur-kofte-cigli-atasehir-mah-cigli-izmir/",
        "Avane İzmir The Bowl": "https://getir.com/yemek/restoran/the-bowl-best-of-we-love-cigli-atasehir-mah-cigli-izmir",
        "Avane İzmir Veganista": "https://getir.com/yemek/restoran/veganista-cigli-atasehir-mah-cigli-izmir"
     }

    num = str(len(list(url_dict.keys())))
    print(f"{num} restaurant's data are gonna be collected.")

    for url_key in url_dict.keys():
        url = url_dict[url_key]
        url_key_list = url_key.split(" ")
        marka = url_key_list[0]
        url_key_list.pop(0)
        branch = url_key_list[0]
        url_key_list.pop(0)
        brand = " ".join(url_key_list)
        now = datetime.now()
        format_date = "%d/%m/%Y"
        format_hour = "%H:%M:%S"
        # format datetime using strftime()
        date = now.strftime(format_date)
        hour = now.strftime(format_hour)
        r = None

        index = str(list(url_dict.keys()).index(url_key))

        # This is for 'if connection get suspened by getir.com because we are making so much request in a short time'
        try:
            r = requests.get(url)
        except Exception as e:
            log = str(e)

            save_to_excel_log(
                ws=ws,
                wb=wb,
                myFileName=myFileName,
                marka=marka,
                branch=branch,
                brand=brand,
                date=date,
                hour=hour,
                log=log,
            )

            save_to_db_log(branch=branch, brand=brand, date=date,hour=hour, log=log)

            print(index + log)

            continue

        if r == None:
            print(index + "URL is None")
            continue

        # TODO: This is gonna change, because we are not checking the restauant open/close status by status.code normally. # Have the screenshot in the screenshots.

        soup = BeautifulSoup(r.content, "html.parser")

        try:
            close = get_close(soup=soup)

        except Exception as e:
            log = str(e)

            save_to_excel_log(
                ws=ws,
                wb=wb,
                myFileName=myFileName,
                marka=marka,
                branch=branch,
                brand=brand,
                date=date,
                hour=hour,
                log=log,
            )

            save_to_db_log(marka=marka , branch=branch, brand=brand, date=date, hour=hour, log=log)

            print(index + log)
            continue

        try:
            current_rating = get_rating(soup=soup)

        except Exception as e:
            current_rating = None

        # Check that if close state is exsist in the html.
        if close == None:
            # TODO: append the date to the current sheet. But this is gonna change by the restaurant.
            status = "AÇIK"

            try:

                save_to_excel(
                    ws=ws,
                    wb=wb,
                    myFileName=myFileName,
                    marka=marka,
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

                print(index + "Changes saved to excel!")

            except Exception as e:
                print(index + str(e))
                print(index + "Couldn't write to excel! Keep looping")
                continue

            try:
                save_to_db(
                    marka=marka,
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

            except Exception as e:
                print(index + str(e))
                print("Couldn't write to db!")
                continue

        # TODO: This is gonna be canceled because we are not going to calculate the current OPEN/CLOSE status with this
        else:
            status = "KAPALI"

            try:

                save_to_excel(
                    ws=ws,
                    wb=wb,
                    myFileName=myFileName,
                    marka=marka,
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

                print(index + "Changes saved to excel!")

            except Exception as e:

                print(index + str(e))
                print(index + "Couldn't write to excel! Keep looping")
                continue

            try:
                save_to_db(
                    marka=marka,
                    branch=branch,
                    brand=brand,
                    date=date,
                    hour=hour,
                    status=status,
                    current_rating=current_rating,
                )

            except Exception as e:
                print(index + str(e))
                print("Couldn't write to db!")
                continue

        sleep(5)

    return url_dict  # temp


def get_rating(soup):
    # Find the rate
    body = soup.body
    s = body.find("div", id="__next")
    s2 = s.find("div", class_="sc-212542e0-2 ckZpLq")
    s4 = s2.find("main", class_="sc-212542e0-0 iiapCb")
    s5 = s4.find("div", class_="sc-e85e5299-0 sc-4e0754cc-0 hkaVQN klYfLJ")
    s6 = s5.find("div", class_="sc-4e0754cc-1 YPsgm")
    s7 = s6.find("div", class_="sc-4e0754cc-3 bwTzrw")
    s8 = s7.find("div", class_="sc-7047f3e2-6 iFptQI")
    s9 = s8.find("div", class_="style__Wrapper-sc-__sc-sbxwka-15 jPuQcd")
    s10 = s9.find("div", class_="style__CardWrapper-sc-__sc-sbxwka-12 ccsSiU")
    s11 = s10.find("div", class_="style__ContentWrapper-sc-__sc-sbxwka-7 emAjmS")
    s12 = s11.find("div", class_="sc-7047f3e2-0 iJHJBI")
    s13 = s12.find("div", class_="sc-7047f3e2-3 hbiBbV")
    rate = s13.find(
        "span", class_="style__Text-sc-__sc-1nwjacj-0 jbOUDC sc-7047f3e2-8 iFDpNz"
    )

    current_rating = rate.get_text()

    return current_rating


def get_close(soup):
    body = soup.body
    s11 = body.find("div", id="__next")
    s10 = s11.find("div", class_="sc-212542e0-2 ckZpLq")
    s9 = s10.find("main", class_="sc-212542e0-0 iiapCb")
    s8 = s9.find("div", class_="sc-e85e5299-0 sc-4e0754cc-0 hkaVQN klYfLJ")
    s7 = s8.find("div", class_="sc-4e0754cc-1 YPsgm")
    s6 = s7.find("div", class_="sc-4e0754cc-3 bwTzrw")
    s5 = s6.find("div", class_="sc-7047f3e2-6 iFptQI")
    s4 = s5.find("div", class_="style__Wrapper-sc-__sc-sbxwka-15 jPuQcd")
    s3 = s4.find("div", class_="style__CardWrapper-sc-__sc-sbxwka-12 ccsSiU")
    s2 = s3.find("div", class_="style__ContentWrapper-sc-__sc-sbxwka-7 emAjmS")
    close = s2.find("div", class_="sc-e27f3f42-0 hPdSRl")

    return close


@repeat(every(1).hour)  # .until("18:30")
def getir_to_excel():

    getir_to_excel_first()


# for the first time and then let the schdule does it's job
getir_to_excel_first()

while True:
    run_pending()
    sleep(5)
